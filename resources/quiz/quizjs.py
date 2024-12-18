# quiz.js を生成するための Python スクリプト
import os
import csv
import openpyxl

debug_mode = False

# 古い資産を削除
js_path = os.path.dirname(__file__) + '/quiz.js'

if os.path.exists(js_path):
    os.remove(js_path) 

# quizjs.xlsx から問題を読み込む
wb = openpyxl.load_workbook(os.path.dirname(__file__) + '/quizjs.xlsx', data_only=True)
# シートを取得
sheet = wb["質問リスト"]
# シート内のセルを取得
rows = sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=6, values_only=True)
# データをリストに格納
data = []
for row in rows:
    data.append(row)

# 設定シートから問題数とカウントダウン時間を取得
sheet = wb["設定"]
numOfQuiz = sheet['B1'].value
countsPerQuestion = sheet['B2'].value

# excelを閉じる
wb.close()

# data を更にカンマ区切りでリスト化する
data = [list(d) for d in data]

scripts = '''
document.addEventListener("DOMContentLoaded", function () {
    const questions = [
'''

# ルビ表記関数
def ruby_text(text):
    text = text.replace("肺", "<ruby>肺<rt> はい</rt></ruby>")
    text = text.replace("胃", "<ruby>胃<rt> い</rt></ruby>")
    text = text.replace("腎臓", "<ruby>腎臓<rt> じんぞう</rt></ruby>")
    text = text.replace("膵臓", "<ruby>膵臓<rt> すいぞう</rt></ruby>")
    text = text.replace("脾臓", "<ruby>脾臓<rt> ひぞう</rt></ruby>")
    text = text.replace("肝臓", "<ruby>肝臓<rt> かんぞう</rt></ruby>")
    text = text.replace("心臓", "<ruby>心臓<rt> しんぞう</rt></ruby>")
    text = text.replace("胆嚢", "<ruby>胆嚢<rt> たんのう</rt></ruby>")
    text = text.replace("小腸", "<ruby>小腸<rt> しょうちょう</rt></ruby>")
    text = text.replace("大腸", "<ruby>大腸<rt> だいちょう</rt></ruby>")
    text = text.replace("食道", "<ruby>食道<rt> しょくどう</rt></ruby>")
    text = text.replace("気道", "<ruby>気道<rt> きどう</rt></ruby>")
    text = text.replace("公道", "<ruby>公道<rt> こうどう</rt></ruby>")
    return text

# Excel から取得したデータを整形して scripts に追加
row_num = 0
index_array = ['question: ', 'images: ', 'options: ', 'answer: ', 'explanation: ']
question_data = ""
for d in data:
    if row_num == 0:
        question_data += "        {\n"

    if d[2] is not None:
        if row_num == 0 or row_num == 2 or row_num == 4:
            d[2] = ruby_text(d[2])
            
        question_data += "            " + index_array[row_num] + d[2] + ",\n"

    if row_num == 4:
        question_data += "        },\n"
        row_num = 0
    else:
        row_num += 1

# 最後の1文字を削除
if question_data.endswith(",\n"):
    question_data = question_data[:-2] + "\n"

scripts += question_data

scripts += f'''
    ];
    const chrsL = ["images/chr1.png", "images/chr2.png", "images/chr3.png", "images/chr4.png", "images/chr5.png", "images/chr6.png", "images/chr7.png", "images/chr8.png"];
    const chrsR = ["images/chr1.png", "images/chr2.png", "images/chr3.png", "images/chr4.png", "images/chr5.png", "images/chr6.png", "images/chr7.png", "images/chr8.png"];

    let countsPerQuestion = {countsPerQuestion};
    const numOfQuiz = {numOfQuiz};
'''

scripts += '''
    let currentQuestionIndex = 0;
    let currentOptionName = "";
    let currentOptionButton = null;
    let score = 0;
    let isInit = true;
    let username = "名無し";
    let selectedOptions = [];
    let answerResults = [];

    const questionContainer = document.getElementById("questionContainer");
    const optionsContainer = document.getElementById("optionsContainer");
    const nextButton = document.getElementById("nextButton");
    const scoreDisplay = document.getElementById("score");
    const answersContainer = document.getElementById("answersContainer");
    const progressBar = document.getElementById("progress");
    // const imagesContainer = document.getElementById("imagesContainer");
    const okSound = new Audio("sound/ok.mp3");
    const ngSound = new Audio("sound/ng.mp3");
    const nameinputContainer = document.getElementById("nameinput-container");
    const quizContainer = document.getElementById("quiz-container");
    const sleep = (time) => new Promise((resolve) => setTimeout(resolve, time));//timeはミリ秒
    const answerOkContainer = document.getElementById("answerOkContainer");
    const answerNGContainer = document.getElementById("answerNGContainer");
    const timeoverContainer = document.getElementById("timeoverContainer");

    // 問題順をシャッフルするための関数
    function shuffle(array) {
        for (let i = array.length - 1; i > 0; i--) {
            const j = Math.floor(Math.random() * (i + 1));
            [array[i], array[j]] = [array[j], array[i]];
        }
    }

    // クイズの進捗具合をプログレスバーで表示するための関数
    function updateProgressBar() {
        var progress;
        if (currentQuestionIndex == numOfQuiz) {
            progress = ((currentQuestionIndex + 1) / numOfQuiz) * 100;
        } else {
            progress = ((currentQuestionIndex) / numOfQuiz) * 100;
        }
        progressBar.style.width = `${progress}%`;
    }

    // クイズを表示する関数
    function displayQuestion() {
        const currentQuestion = questions[currentQuestionIndex];
        questionContainer.innerHTML = currentQuestion.question;
        // 左側に表示するキャラクターの定義
        document.getElementById("leftchr").innerHTML = "";
        const imgElementL = document.createElement("img");
        imgElementL.src = chrsL[currentQuestionIndex];
        imgElementL.classList.add("img-fluid", "mb-3");
        imgElementL.style.height = "160px";
        document.getElementById("leftchr").appendChild(imgElementL);
        // 右側に表示するキャラクターの定義
        document.getElementById("rightchr").innerHTML = "";
        const imgElementR = document.createElement("img");
        imgElementR.src = chrsR[currentQuestionIndex];
        imgElementR.classList.add("img-fluid", "mb-3");
        imgElementR.style.height = "160px";
        document.getElementById("rightchr").appendChild(imgElementR);

        // 画像一覧を生成
        // imagesContainer.innerHTML = "";
        // currentQuestion.images.forEach((image) => {
        //     const imageElement = document.createElement("img");
        //     imageElement.src = image;
        //     imageElement.classList.add("img-fluid", "mb-3");
        //     imageElement.style.width = "200px";
        //     imageElement.style.height = "200px";
        //     imagesContainer.appendChild(imageElement);
        // });

        // 選択肢のボタンを生成
        optionsContainer.innerHTML = "";
        i = 0;
        var numOption = currentQuestion.options.length;
        currentQuestion.options.forEach((option) => {
            const optionButton = document.createElement("button");
            optionButton.classList.add("btn", "btn-outline-primary", "option");
            optionButton.innerHTML = option;
            optionButton.addEventListener("click", () =>
                selectOption(optionButton, option)
            );
            optionsContainer.appendChild(optionButton);
            if(i < numOption - 1) {
                const spacer = document.createElement("span");
                spacer.style.height = "10px";
                spacer.style.width = "10px";
                spacer.innerHTML = "&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;";
                optionsContainer.appendChild(spacer);
            }
            i++;
        });

        // 現在の問題番号をプログレスバーに反映
        updateProgressBar();
        nextButton.disabled = true;

        // 現在の問題番号を表示する
        document.getElementById("questionNumber").textContent = `問題 ${currentQuestionIndex + 1} / ${numOfQuiz}`;
        // 問題開始とともにカウントダウンを開始
        clearInterval(countdownInterval);
        setRedirectInfo("#", countsPerQuestion);
    }

    // 選択肢をクリックしたときの処理
    function selectOption(optionButton, selectedOption) {
        // 選択している選択肢の色を変更
        const optionButtons = document.querySelectorAll(".option");
        optionButtons.forEach((button) => {
            button.classList.remove("optionSelected");
            button.classList.add("btn", "btn-outline-primary", "option");
        });
        optionButton.classList.add("optionSelected");
        currentOptionName = optionButton.textContent;
        currentOptionButton = optionButton;
        nextButton.disabled = false;
    }

    // 次へボタンを押したときの処理
    nextButton.addEventListener("click", async () => {
        currentOptionButton.classList.remove("optionSelected","btn-outline-primary");
        nextButton.style.visibility = "hidden";
        selectedOptions.push(currentOptionName);
        if(document.getElementById("remainTime").innerText == 0) {
            // 時間切れのためスコア加算なし
            currentOptionButton.classList.add("optionNG");
            timeoverContainer.style.display = "block";
            await sleep(3000);
            timeoverContainer.style.display = "none";
            answerResults.push("時間切れ");
        }else if(currentOptionName === questions[currentQuestionIndex].answer) {
            okSound.currentTime = 0;
            okSound.play();
            score += 100 / numOfQuiz;
            currentOptionButton.classList.add("optionOK");
            answerOkContainer.style.display = "block";
            await sleep(3000);
            answerOkContainer.style.display = "none";
            answerResults.push("○");
        }else {
            ngSound.currentTime = 0;
            ngSound.play();
            currentOptionButton.classList.add("optionNG");
            answerNGContainer.style.display = "block";
            await sleep(3000);
            answerNGContainer.style.display = "none";
            answerResults.push("×");
        }
        scoreDisplay.textContent = `${username}さんのスコア: ${Math.round(score)}`;

        currentQuestionIndex++;
        if (currentQuestionIndex < numOfQuiz) {
            document.getElementById("countContainer").style.display = "block";
            nextButton.style.visibility = "visible";
            displayQuestion();
        } else {
            displayScore();
            displayAnswers();
            updateProgressBar();
        }
        document.getElementById("messageContainer").innerHTML = "";
        document.getElementById("remainTime").style.color = "black";
    });

    // 成績を表示する
    function displayScore() {
        questionContainer.textContent = "";
        optionsContainer.innerHTML = "";
        nextButton.style.display = "none";
        scoreDisplay.textContent = `${username}さんのスコアは ${Math.round(score)} / 100 です。`;
        document.getElementById("countContainer").style.display = "none";
        // document.getElementById("imagesContainer").style.display = "none";
        document.getElementById("questionNumber").textContent = "";
        document.getElementById("qrContainer").style.display = "none";
        document.getElementById("leftchr").style.display = "none";
        document.getElementById("rightchr").style.display = "none";
    }

    // 解説を表示する
    function displayAnswers() {
        answersContainer.innerHTML = "<h4>解答と解説</h4>";
        var tempStr = "";
        var tempAns = "";
        var tempSel = "";
        questions.forEach((question, index) => {
            const answerDiv = document.createElement("div");
            answerDiv.classList.add("mb-2");
            tempAns = question.answer.split(" ")[0];
            tempSel = selectedOptions[index] + "";
            tempSel = tempSel.split(" ")[0];
            if(answerResults[index] == "○") {
                tempStr = "";
            }else {
                tempStr = `(正解は ${tempAns})`;
            }
            answerDiv.innerHTML = `<strong>質問 ${index + 1}:</strong> ${question.question}
                                   ≪
                                   <strong>$選択:</strong> ${tempSel}
                                   <strong>[${answerResults[index]}]</strong>
                                   ${tempStr}
                                   ≫<br>
                                   <strong>解説:</strong> ${question.explanation}
                                   `;
            if(index < numOfQuiz) {
                answersContainer.appendChild(answerDiv);
            }
        });
        document.getElementById("questionNumber").innerHTML = `
                <img src="images/chr7.png" height="30px" />&nbsp;
                <img src="images/chr5.png" height="30px" />&nbsp;
                お疲れ様でした！
                &nbsp;<img src="images/chr3.png" height="30px" />
                &nbsp;<img src="images/chr1.png" height="30px" />
                `;
        document.getElementById("remainTime").textContent = "";
    }

    // 初回処理 (全体問題のシャッフルと最初の問題の表示)
    shuffle(questions);
    shuffle(chrsL);
    shuffle(chrsR);
    displayQuestion();
    answerOkContainer.style.display = "none";
    answerNGContainer.style.display = "none";
    timeoverContainer.style.display = "none";

    if(isInit) {
        document.getElementById("startQuizButton").addEventListener("click", () =>
            startQuiz()
        );
        quizContainer.style.display = "none";
        nameinputContainer.style.display = "block";
    }else {
        quizContainer.style.display = "block";
        nameinputContainer.style.display = "none";
    }    

    // クイズをはじめるボタンクリック時の処理
    function startQuiz() {
        if (document.getElementById("username").value == "") {
            alert("名前を入力してください");
            return;
        }
        isInit = false;
        quizContainer.style.display = "block";
        nameinputContainer.style.display = "none";
        username = document.getElementById("username").value;
        clearInterval(countdownInterval);
        setRedirectInfo("#", countsPerQuestion);
    }
});

// カウントダウン関数はグローバル領域に定義する
let countdownInterval;

function setRedirectInfo(url, time) {
    if (url && time) {
        var countdown = time;
        document.getElementById("remainTime").innerText = countdown;

        countdownInterval = setInterval(function() {
            countdown--;
            document.getElementById("remainTime").innerText = countdown;

            if (countdown <= 10 &&
                    document.getElementById("quiz-container").style.display != "none") {
                document.getElementById("remainTime").style.color = "red";
            }

            if (countdown <= 0) {
                clearInterval(countdownInterval);
                if (document.getElementById("countContainer").style.display != "none" &&
                        document.getElementById("quiz-container").style.display != "none") {
                    document.getElementById("messageContainer").innerHTML = "時間切れです。";
                }
            }
        }, 1000); // 1秒ごとにカウントダウン
    } else {
        alert("正しいURLと時間を入力してください");
    }
}
'''

with open(js_path, 'a', encoding='utf-8') as f:
    f.write(scripts)
